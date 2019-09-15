'''
@author: Administrator
@Wechat Contact:第一行Python代码 
@project: python21
@file: class_day14.py
@time: 2019/8/21 13:05
@desc:
'''
import openpyxl
class read_excel(object):
    def __init__(self,workbook,sheet):
        self.workbook=workbook
        self.sheet=sheet
        super().__init__(self)
    def open_wookbook(self):
        #打开工作部
        self.wb =openpyxl.load_workbook(self.workbook)
        #打开表单
        self.sh=self.wb[self.sheet]
    #读取单元格数据
    # c11 =sheet.cell(row=1,column=1).value
    # print(c11)
    #循环读取表格数据

    def read_excel(self):
    # cases = [
    #     {"excepted":{"code": 1, "msg": "注册成功"},"data":('python21', '1234567', '1234567')},
    #     {"excepted":{"code": 0, "msg": "两次密码不一致"},"data":('python22','1234567','12334567')},
    #     {"excepted":{"code": 0, "msg": "两次密码不一致"},"data":('python22','12334567','1234567')}
    # ]
        self.open_wookbook()
        max_row=self.sh.max_rows
        max_column=self.sh.max_columns
        for row1 in range(1,max_row+1):
            list1=[]
            if row1== 1:
                titles1=[]
                for column1 in range(1,4):
                    data = self.sh.cell(row=row1,column=column1).value
                    # print("-----------------------")
                    # print(data)
                    # print("-----------------------")
                    titles1.append(data)
                #print(titles1)
            else:
                datas=[]
                for column1 in range(1,max_column+1):
                    data = self.sh.cell(row=row1,column=column1).value
                    datas.append(data)
                #print(datas)
                #打包
                print("-----------------------")
                cases=dict(zip(titles1,datas))
        return  cases
if __name__=='__main':
    cases =read_excel('cases.xlsx','Sheet2').read_excel()
    print(cases)